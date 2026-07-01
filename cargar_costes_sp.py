"""
cargar_costes_sp.py  —  v2 (unificado)
───────────────────────────────────────
Actualiza DOS listas de SharePoint con los valores económicos del Excel Maestro:

  1. M_Partidas (Producción)
     Escribe: CosteUd (col K) y CosteTot (col AC)

  2. M_PartidasControl (Control)
     Escribe: field_4 (PrecioUd), field_5 (CosteUd),
              field_6 (VentaTot), field_7 (CosteTot)

  Procesa los dos Excels Maestro de Obra en una sola ejecución:
    - 23-039_Maestro_Obra_13_DEF.xlsx
    - 26-008_Maestro_Obra_DEF.xlsx

PREREQUISITOS:
  pip install openpyxl msal requests --break-system-packages

CONFIGURACIÓN: rellenar las 4 constantes del bloque CONFIG antes de ejecutar.
"""

import json, sys, time
from openpyxl import load_workbook

# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIG — rellenar antes de ejecutar
# ═══════════════════════════════════════════════════════════════════════════════
CLIENT_ID  = "ce963b14-1236-4a40-a127-3a055e964cff"
TENANT_ID  = "ffc698d1-4a6a-46f0-ad29-64cf6b17ab22"
SITE_URL   = "https://telice.sharepoint.com/sites/APPS_TEST_1"
EXCEL_FILES = [
    "23-039_Maestro_Obra_13_DEF.xlsx",
    "26-008_Maestro_Obra_DEF.xlsx",
]
# ═══════════════════════════════════════════════════════════════════════════════

# Definición de las dos listas a actualizar
LISTAS = [
    {
        "nombre":       "M_Partidas",           # nombre exacto en SP
        "hoja":         "Partidas_Produccion",  # hoja del Excel
        "fila_inicio":  5,                      # primera fila con datos (1-indexed)
        "col_codigo":   0,                      # A → Title en SP
        "col_medicion": 8,                      # I → para filtrar filas vacías
        "campos": {                             # campo_sp: índice_columna_excel (0-indexed)
            "CosteUd":  10,   # K = Coste UD Proyecto
            "CosteTot": 28,   # AC = TOTAL Coste
        },
        "excluir": [],        # valores de código a ignorar
        "decimales": {
            "CosteUd":  6,
            "CosteTot": 2,
        }
    },
    {
        "nombre":       "M_PartidasControl",    # ajustar si el nombre en SP difiere
        "hoja":         "Partidas_Control",
        "fila_inicio":  3,
        "col_codigo":   0,                      # A → Title en SP
        "col_medicion": 3,                      # D = Medición → para filtrar filas vacías
        "campos": {
            "field_4":  4,    # E = Precio venta €/ud
            "field_5":  5,    # F = Coste estudio €/ud
            "field_6":  6,    # G = Importe venta total
            "field_7":  7,    # H = Coste total estudio
        },
        "excluir": ["TOTAL GENERAL"],
        "decimales": {
            "field_4": 6,
            "field_5": 6,
            "field_6": 2,
            "field_7": 2,
        }
    },
]

# ─── Dependencias ─────────────────────────────────────────────────────────────
try:
    import msal, requests
except ImportError:
    print("ERROR: pip install msal requests --break-system-packages")
    sys.exit(1)

# ─── Autenticación ────────────────────────────────────────────────────────────
def get_token():
    app = msal.PublicClientApplication(
        CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}"
    )
    scopes = ["https://graph.microsoft.com/Sites.ReadWrite.All"]
    accounts = app.get_accounts()
    if accounts:
        r = app.acquire_token_silent(scopes, account=accounts[0])
        if r and "access_token" in r:
            print("✓ Token desde caché")
            return r["access_token"]
    flow = app.initiate_device_flow(scopes=scopes)
    if "user_code" not in flow:
        raise ValueError(f"Error device flow: {flow}")
    print("\n" + "="*60)
    print("AUTENTICACIÓN — abre el navegador y entra el código:")
    print(f"  URL:    {flow['verification_uri']}")
    print(f"  Código: {flow['user_code']}")
    print("="*60 + "\n")
    r = app.acquire_token_by_device_flow(flow)
    if "access_token" not in r:
        raise ValueError(r.get("error_description", r))
    print("✓ Autenticación correcta\n")
    return r["access_token"]

# ─── Graph helpers ────────────────────────────────────────────────────────────
def g(token, url):
    r = requests.get(url, headers={"Authorization": f"Bearer {token}"})
    r.raise_for_status()
    return r.json()

def patch(token, url, body):
    r = requests.patch(
        url,
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
        data=json.dumps(body)
    )
    if r.status_code not in (200, 204):
        raise ValueError(f"{r.status_code}: {r.text[:200]}")

def get_site_id(token):
    parts  = SITE_URL.replace("https://", "").split("/", 1)
    host   = parts[0]
    path   = parts[1] if len(parts) > 1 else ""
    data   = g(token, f"https://graph.microsoft.com/v1.0/sites/{host}:/{path}")
    sid    = data["id"]
    print(f"✓ Site ID: {sid}")
    return sid

def get_lista_id(token, site_id, nombre):
    nombre_enc = requests.utils.quote(nombre)
    data = g(token, f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists?$filter=displayName eq '{nombre}'")
    items = data.get("value", [])
    if not items:
        raise ValueError(f"Lista '{nombre}' no encontrada. Verifica el nombre exacto en SP.")
    lid = items[0]["id"]
    print(f"✓ Lista '{nombre}' → {lid}")
    return lid

def get_all_items(token, site_id, lista_id, nombre):
    url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{lista_id}/items?$select=id,fields/Title&$top=500"
    items = []
    while url:
        data = g(token, url)
        items.extend(data.get("value", []))
        url = data.get("@odata.nextLink")
    print(f"✓ {len(items)} items en '{nombre}'")
    return {item["fields"]["Title"]: item["id"] for item in items}

# ─── Lectura Excel ────────────────────────────────────────────────────────────
def leer_excel(cfg):
    wb   = load_workbook(cfg.get("excel", EXCEL_FILES[0]), read_only=True, data_only=True)
    ws   = wb[cfg["hoja"]]
    cols = cfg["campos"]
    dec  = cfg["decimales"]
    excl = set(cfg.get("excluir", []))
    datos = {}
    vacias = 0

    for row in ws.iter_rows(min_row=cfg["fila_inicio"], values_only=True):
        cod = row[cfg["col_codigo"]]
        if not cod or not isinstance(cod, str):
            vacias += 1
            continue
        cod = cod.strip()
        if cod in excl:
            continue
        med = row[cfg["col_medicion"]]
        if not med:
            vacias += 1
            continue
        entrada = {}
        for campo_sp, col_idx in cols.items():
            val = row[col_idx]
            d   = dec.get(campo_sp, 2)
            entrada[campo_sp] = round(float(val), d) if val is not None else 0.0
        datos[cod] = entrada

    wb.close()
    print(f"  {len(datos)} partidas leídas ({vacias} filas vacías/título ignoradas)")
    return datos

# ─── Procesado de una lista ───────────────────────────────────────────────────
def procesar_lista(token, site_id, cfg):
    nombre = cfg["nombre"]
    print(f"\n{'─'*60}")
    print(f"Procesando: {nombre}  ({cfg['hoja']})")

    datos    = leer_excel(cfg)
    lista_id = get_lista_id(token, site_id, nombre)
    mapa_sp  = get_all_items(token, site_id, lista_id, nombre)

    base_url     = f"https://graph.microsoft.com/v1.0/sites/{site_id}/lists/{lista_id}/items"
    actualizados = 0
    no_enc       = []
    errores      = []
    total        = len(datos)

    print(f"  Actualizando {total} partidas...")
    for i, (cod, campos) in enumerate(datos.items(), 1):
        if cod not in mapa_sp:
            no_enc.append(cod)
            continue
        try:
            patch(token, f"{base_url}/{mapa_sp[cod]}", {"fields": campos})
            actualizados += 1
            if i % 25 == 0 or i == total:
                print(f"    {i}/{total} — {actualizados} OK, {len(errores)} errores")
            time.sleep(0.05)   # evitar throttling Graph API
        except Exception as e:
            errores.append(f"{cod}: {e}")

    # Resumen de lista
    print(f"\n  RESULTADO '{nombre}':")
    print(f"    ✓ Actualizados:   {actualizados}")
    print(f"    ⚠ No en SP:       {len(no_enc)}")
    print(f"    ✗ Errores:        {len(errores)}")
    if no_enc:
        print(f"    Códigos no encontrados (primeros 5): {no_enc[:5]}")
    if errores:
        print(f"    Errores (primeros 3): {errores[:3]}")

    return {"actualizados": actualizados, "no_enc": len(no_enc), "errores": len(errores)}

# ─── Main ─────────────────────────────────────────────────────────────────────
def main():
    print("╔══════════════════════════════════════════════════════════╗")
    print("║  Carga de costes unitarios y totales → SharePoint        ║")
    print("║  Listas: M_Partidas + M_PartidasControl                  ║")
    print("╚══════════════════════════════════════════════════════════╝\n")

    # Validar config básica
    if "TU_" in CLIENT_ID or "TU_" in TENANT_ID:
        print("ERROR: Rellena CLIENT_ID y TENANT_ID en la sección CONFIG del script.")
        sys.exit(1)

    token   = get_token()
    site_id = get_site_id(token)

    totales = {"actualizados": 0, "no_enc": 0, "errores": 0}
    for excel in EXCEL_FILES:
        print(f"\n{'═'*60}")
        print(f"EXCEL: {excel}")
        for cfg in LISTAS:
            cfg_obra = dict(cfg)
            cfg_obra["excel"] = excel
            res = procesar_lista(token, site_id, cfg_obra)
            for k in totales:
                totales[k] += res[k]

    print(f"\n{'═'*60}")
    print("RESUMEN GLOBAL")
    print(f"  ✓ Actualizados totales: {totales['actualizados']}")
    print(f"  ⚠ No encontrados:       {totales['no_enc']}")
    print(f"  ✗ Errores:              {totales['errores']}")
    print("Done.")

if __name__ == "__main__":
    main()
